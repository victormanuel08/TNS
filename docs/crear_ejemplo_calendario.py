"""
Script para crear un Excel de ejemplo para carga masiva del calendario tributario
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date

# Crear workbook
wb = openpyxl.Workbook()

# Eliminar hoja por defecto
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ========== HOJA 1: INSTRUCCIONES ==========
ws_inst = wb.create_sheet("INSTRUCCIONES", 0)
ws_inst['A1'] = "INSTRUCCIONES PARA CARGA MASIVA DE CALENDARIO TRIBUTARIO"
ws_inst['A1'].font = Font(bold=True, size=14)
ws_inst['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_inst['A1'].font = Font(bold=True, size=14, color="FFFFFF")

ws_inst['A3'] = "Columnas requeridas:"
ws_inst['A3'].font = Font(bold=True)
ws_inst['A4'] = "• tax_code: Código del impuesto (RGC, RPJ, RPN, IVB, IVC, AEE, RSA, RET, etc.)"
ws_inst['A5'] = "• expirations_digits: Últimos 1 o 2 dígitos del NIT ('1'-'9', '0', '01'-'99', '00', o '' para todos)"
ws_inst['A6'] = "• third_type_code: Tipo de tercero ('PN' = Persona Natural, 'PJ' = Persona Jurídica, o '' para todos)"
ws_inst['A7'] = "• regiment_type_code: Régimen tributario ('GC' = Gran Contribuyente, 'SIM' = Simple, 'ORD' = Ordinario, o '' para todos)"
ws_inst['A8'] = "• date: Fecha límite en formato DD/MM/YYYY o YYYY-MM-DD"
ws_inst['A9'] = "• description: Descripción de la obligación tributaria"

ws_inst['A11'] = "NOTAS IMPORTANTES:"
ws_inst['A11'].font = Font(bold=True)
ws_inst['A12'] = "1. El tax_code debe existir previamente en el sistema"
ws_inst['A13'] = "2. Si expirations_digits está vacío (''), aplica a TODOS los NITs"
ws_inst['A14'] = "3. Si third_type_code está vacío, aplica a todos los tipos de tercero"
ws_inst['A15'] = "4. Si regiment_type_code está vacío, aplica a todos los regímenes"
ws_inst['A16'] = "5. Las fechas se pueden repetir para diferentes combinaciones de dígitos"

# Ajustar ancho de columnas
ws_inst.column_dimensions['A'].width = 100

# ========== HOJA 2: CALENDARIO TRIBUTARIO (EJEMPLO) ==========
ws_cal = wb.create_sheet("CALENDARIO_TRIBUTARIO", 1)

# Encabezados
headers = ['tax_code', 'expirations_digits', 'third_type_code', 'regiment_type_code', 'date', 'description']
for col, header in enumerate(headers, 1):
    cell = ws_cal.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Datos de ejemplo (basados en create_expirations.py de BCE)
ejemplos = [
    # RGC - Gran Contribuyente, Persona Natural
    {'tax_code': 'RGC', 'expirations_digits': '1', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '11/02/2025', 'description': 'Pago primera cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '2', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '12/02/2025', 'description': 'Pago primera cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '3', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '13/02/2025', 'description': 'Pago primera cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '0', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '24/02/2025', 'description': 'Pago primera cuota'},
    
    # RGC - Segunda cuota
    {'tax_code': 'RGC', 'expirations_digits': '1', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '09/04/2025', 'description': 'Declaración y pago segunda cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '2', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '10/04/2025', 'description': 'Declaración y pago segunda cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '0', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '24/04/2025', 'description': 'Declaración y pago segunda cuota'},
    
    # RGC - Persona Jurídica
    {'tax_code': 'RGC', 'expirations_digits': '1', 'third_type_code': 'PJ', 'regiment_type_code': 'GC', 'date': '11/02/2025', 'description': 'Pago primera cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '2', 'third_type_code': 'PJ', 'regiment_type_code': 'GC', 'date': '12/02/2025', 'description': 'Pago primera cuota'},
    {'tax_code': 'RGC', 'expirations_digits': '0', 'third_type_code': 'PJ', 'regiment_type_code': 'GC', 'date': '24/02/2025', 'description': 'Pago primera cuota'},
    
    # RPJ - Persona Jurídica, sin régimen específico
    {'tax_code': 'RPJ', 'expirations_digits': '1', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '12/05/2025', 'description': 'Declaración y pago primera cuota'},
    {'tax_code': 'RPJ', 'expirations_digits': '2', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '13/05/2025', 'description': 'Declaración y pago primera cuota'},
    {'tax_code': 'RPJ', 'expirations_digits': '0', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '23/05/2025', 'description': 'Declaración y pago primera cuota'},
    
    # RPN - Persona Natural, con últimos 2 dígitos
    {'tax_code': 'RPN', 'expirations_digits': '01', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '12/08/2025', 'description': 'Declaracion y Pago'},
    {'tax_code': 'RPN', 'expirations_digits': '02', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '12/08/2025', 'description': 'Declaracion y Pago'},
    {'tax_code': 'RPN', 'expirations_digits': '10', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '19/08/2025', 'description': 'Declaracion y Pago'},
    {'tax_code': 'RPN', 'expirations_digits': '50', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '16/09/2025', 'description': 'Declaracion y Pago'},
    {'tax_code': 'RPN', 'expirations_digits': '99', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '24/10/2025', 'description': 'Declaracion y Pago'},
    {'tax_code': 'RPN', 'expirations_digits': '00', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '24/10/2025', 'description': 'Declaracion y Pago'},
    
    # IVB - IVA Bimestral
    {'tax_code': 'IVB', 'expirations_digits': '1', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '15/03/2025', 'description': 'Declaración IVA Bimestral'},
    {'tax_code': 'IVB', 'expirations_digits': '2', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '16/03/2025', 'description': 'Declaración IVA Bimestral'},
    {'tax_code': 'IVB', 'expirations_digits': '0', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '20/03/2025', 'description': 'Declaración IVA Bimestral'},
    
    # IVC - IVA Cuatrimestral
    {'tax_code': 'IVC', 'expirations_digits': '1', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '15/03/2025', 'description': 'Declaración IVA Cuatrimestral'},
    {'tax_code': 'IVC', 'expirations_digits': '2', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '16/03/2025', 'description': 'Declaración IVA Cuatrimestral'},
    
    # AEE - Otros impuestos, Gran Contribuyente
    {'tax_code': 'AEE', 'expirations_digits': '1', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '09/04/2025', 'description': ''},
    {'tax_code': 'AEE', 'expirations_digits': '2', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '10/04/2025', 'description': ''},
    {'tax_code': 'AEE', 'expirations_digits': '0', 'third_type_code': 'PN', 'regiment_type_code': 'GC', 'date': '24/04/2025', 'description': ''},
    
    # AEE - Con últimos 2 dígitos
    {'tax_code': 'AEE', 'expirations_digits': '01', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '12/08/2025', 'description': ''},
    {'tax_code': 'AEE', 'expirations_digits': '50', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '16/09/2025', 'description': ''},
    {'tax_code': 'AEE', 'expirations_digits': '99', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '24/10/2025', 'description': ''},
    
    # RSA - Declaración Anual
    {'tax_code': 'RSA', 'expirations_digits': '1', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '15/04/2025', 'description': 'DECLARACION ANUAL AÑO GRAVABLE'},
    {'tax_code': 'RSA', 'expirations_digits': '2', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '15/04/2025', 'description': 'DECLARACION ANUAL AÑO GRAVABLE'},
    {'tax_code': 'RSA', 'expirations_digits': '0', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '23/04/2025', 'description': 'DECLARACION ANUAL AÑO GRAVABLE'},
    
    # RET - Retención mensual
    {'tax_code': 'RET', 'expirations_digits': '1', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '11/02/2025', 'description': 'ENERO'},
    {'tax_code': 'RET', 'expirations_digits': '2', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '12/02/2025', 'description': 'ENERO'},
    {'tax_code': 'RET', 'expirations_digits': '0', 'third_type_code': 'PJ', 'regiment_type_code': '', 'date': '20/02/2025', 'description': 'ENERO'},
    
    # Ejemplo con dígitos vacíos (aplica a todos)
    {'tax_code': 'AEE', 'expirations_digits': '', 'third_type_code': 'PN', 'regiment_type_code': '', 'date': '31/12/2025', 'description': 'Obligación general para todos los NITs'},
]

# Agregar datos
for row_idx, ejemplo in enumerate(ejemplos, 2):
    ws_cal.cell(row=row_idx, column=1, value=ejemplo['tax_code'])
    ws_cal.cell(row=row_idx, column=2, value=ejemplo['expirations_digits'])
    ws_cal.cell(row=row_idx, column=3, value=ejemplo['third_type_code'])
    ws_cal.cell(row=row_idx, column=4, value=ejemplo['regiment_type_code'])
    ws_cal.cell(row=row_idx, column=5, value=ejemplo['date'])
    ws_cal.cell(row=row_idx, column=6, value=ejemplo['description'])

# Ajustar ancho de columnas
ws_cal.column_dimensions['A'].width = 15  # tax_code
ws_cal.column_dimensions['B'].width = 20  # expirations_digits
ws_cal.column_dimensions['C'].width = 18  # third_type_code
ws_cal.column_dimensions['D'].width = 20  # regiment_type_code
ws_cal.column_dimensions['E'].width = 15  # date
ws_cal.column_dimensions['F'].width = 60  # description

# Congelar primera fila
ws_cal.freeze_panes = 'A2'

# ========== HOJA 3: VALORES VÁLIDOS ==========
ws_val = wb.create_sheet("VALORES_VALIDOS", 2)

ws_val['A1'] = "VALORES VÁLIDOS PARA REFERENCIA"
ws_val['A1'].font = Font(bold=True, size=14)
ws_val['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_val['A1'].font = Font(bold=True, size=14, color="FFFFFF")

# tax_code válidos
ws_val['A3'] = "tax_code válidos (ejemplos):"
ws_val['A3'].font = Font(bold=True)
tax_codes = ['RGC', 'RPJ', 'RPN', 'IVB', 'IVC', 'AEE', 'RSA', 'RET', 'ECT']
for idx, code in enumerate(tax_codes, 4):
    ws_val.cell(row=idx, column=1, value=f"  • {code}")

# expirations_digits válidos
ws_val['A14'] = "expirations_digits válidos:"
ws_val['A14'].font = Font(bold=True)
ws_val['A15'] = "  • '1' a '9' (último dígito)"
ws_val['A16'] = "  • '0' (último dígito cero)"
ws_val['A17'] = "  • '01' a '99' (últimos dos dígitos)"
ws_val['A18'] = "  • '00' (últimos dos dígitos cero)"
ws_val['A19'] = "  • '' (vacío = aplica a todos los NITs)"

# third_type_code válidos
ws_val['A21'] = "third_type_code válidos:"
ws_val['A21'].font = Font(bold=True)
ws_val['A22'] = "  • 'PN' (Persona Natural)"
ws_val['A23'] = "  • 'PJ' (Persona Jurídica)"
ws_val['A24'] = "  • '' (vacío = aplica a todos los tipos)"

# regiment_type_code válidos
ws_val['A26'] = "regiment_type_code válidos:"
ws_val['A26'].font = Font(bold=True)
ws_val['A27'] = "  • 'GC' (Gran Contribuyente)"
ws_val['A28'] = "  • 'SIM' (Régimen Simple)"
ws_val['A29'] = "  • 'ORD' (Régimen Ordinario)"
ws_val['A30'] = "  • '' (vacío = aplica a todos los regímenes)"

ws_val.column_dimensions['A'].width = 60

# Guardar archivo
output_file = 'docs/ejemplo_calendario_tributario.xlsx'
wb.save(output_file)
print(f"✅ Excel de ejemplo creado: {output_file}")
print(f"   - {len(ejemplos)} registros de ejemplo")
print(f"   - 3 hojas: INSTRUCCIONES, CALENDARIO_TRIBUTARIO, VALORES_VALIDOS")

